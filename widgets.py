# widgets.py
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QTableWidget, QTableWidgetItem, QPushButton,
                             QGroupBox, QProgressBar, QComboBox, QFrame,
                             QHeaderView, QMessageBox, QFileDialog)
from PyQt6.QtCore import Qt, QTimer, QDateTime
from PyQt6.QtGui import QColor, QPainter
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class ReportsWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # Report controls
        control_layout = QHBoxLayout()
        
        self.report_type = QComboBox()
        self.report_type.addItems(["Weekly Report", "Monthly Report"])
        control_layout.addWidget(QLabel("Report Type:"))
        control_layout.addWidget(self.report_type)
        
        generate_btn = QPushButton("Generate Report")
        generate_btn.clicked.connect(self.generate_report)
        control_layout.addWidget(generate_btn)
        
        export_btn = QPushButton("Export Report")
        export_btn.clicked.connect(self.export_report)
        control_layout.addWidget(export_btn)
        
        print_btn = QPushButton("Print Report")
        print_btn.clicked.connect(self.print_report)
        control_layout.addWidget(print_btn)
        
        control_layout.addStretch()
        layout.addLayout(control_layout)
        
        # Chart area
        self.figure = Figure(figsize=(8, 6))
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)
        
        self.setLayout(layout)
        
        # Generate initial report
        self.generate_report()
    
    def generate_report(self):
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        if self.report_type.currentText() == "Weekly Report":
            data = self.get_weekly_data()
            ax.set_title('Weekly Profit/Loss Report')
        else:
            data = self.get_monthly_data()
            ax.set_title('Monthly Profit/Loss Report')
        
        if data:
            dates = [row[0] for row in data]
            profits = [row[1] if row[1] else 0 for row in data]
            
            bars = ax.bar(dates, profits)
            ax.set_xlabel('Date')
            ax.set_ylabel('Profit/Loss ($)')
            ax.tick_params(axis='x', rotation=45)
            
            for i, bar in enumerate(bars):
                if profits[i] >= 0:
                    bar.set_color('#4CAF50')
                else:
                    bar.set_color('#f44336')
        else:
            ax.text(0.5, 0.5, 'No data available for this period', 
                   ha='center', va='center', transform=ax.transAxes)
        
        self.figure.tight_layout()
        self.canvas.draw()
    
    def get_weekly_data(self):
        cursor = self.parent.conn.cursor()
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)
        
        cursor.execute("""
            SELECT i.date, SUM(ii.profit) as daily_profit
            FROM invoices i
            LEFT JOIN invoice_items ii ON i.id = ii.invoice_id
            WHERE i.date >= ? AND i.date <= ?
            GROUP BY i.date
            ORDER BY i.date
        """, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        
        return cursor.fetchall()
    
    def get_monthly_data(self):
        cursor = self.parent.conn.cursor()
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)
        
        cursor.execute("""
            SELECT i.date, SUM(ii.profit) as daily_profit
            FROM invoices i
            LEFT JOIN invoice_items ii ON i.id = ii.invoice_id
            WHERE i.date >= ? AND i.date <= ?
            GROUP BY i.date
            ORDER BY i.date
        """, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        
        return cursor.fetchall()
    
    def export_report(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Report", "", "Excel Files (*.xlsx);;PDF Files (*.pdf)")
        if file_path:
            if file_path.endswith('.xlsx'):
                self.export_to_excel(file_path)
            elif file_path.endswith('.pdf'):
                self.export_to_pdf(file_path)
    
    def export_to_excel(self, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        headers = ['Date', 'Profit/Loss']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        
        data = self.get_weekly_data() if self.report_type.currentText() == "Weekly Report" else self.get_monthly_data()
        for row, (date, profit) in enumerate(data, 2):
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=profit if profit else 0)
        
        wb.save(file_path)
        QMessageBox.information(self, "Success", "Report exported successfully!")
    
    def export_to_pdf(self, file_path):
        self.figure.savefig(file_path, format='pdf', bbox_inches='tight')
        QMessageBox.information(self, "Success", "Report exported successfully!")
    
    def print_report(self):
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        dialog = QPrintDialog(printer, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            painter = QPainter(printer)
            self.figure.savefig(painter, format='pdf')
            painter.end()

class DashboardWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.setup_ui()
        
        # Setup timer for updates
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_dashboard)
        self.timer.start(60000)  # Update every minute
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # Welcome message
        welcome_label = QLabel(f"Welcome to {self.parent.software_name if self.parent else 'Business Software'}")
        welcome_label.setStyleSheet("font-size: 24px; font-weight: bold; margin: 10px; color: #333;")
        welcome_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(welcome_label)
        
        # Stats cards
        stats_layout = QHBoxLayout()
        
        # Sales card
        sales_card = self.create_stat_card("Total Sales", "$0", "#4CAF50")
        self.sales_label = sales_card.findChild(QLabel, "value_label")
        stats_layout.addWidget(sales_card)
        
        # Profit card
        profit_card = self.create_stat_card("Total Profit", "$0", "#2196F3")
        self.profit_label = profit_card.findChild(QLabel, "value_label")
        stats_layout.addWidget(profit_card)
        
        # Products card
        products_card = self.create_stat_card("Total Products", "0", "#FF9800")
        self.products_label = products_card.findChild(QLabel, "value_label")
        stats_layout.addWidget(products_card)
        
        # Invoices card
        invoices_card = self.create_stat_card("Total Invoices", "0", "#9C27B0")
        self.invoices_label = invoices_card.findChild(QLabel, "value_label")
        stats_layout.addWidget(invoices_card)
        
        layout.addLayout(stats_layout)
        
        # Progress bars
        progress_layout = QHBoxLayout()
        
        # Sales progress
        sales_progress_group = QGroupBox("Sales Progress")
        sales_progress_group.setStyleSheet("QGroupBox { font-weight: bold; color: #4CAF50; }")
        sales_progress_layout = QVBoxLayout()
        self.sales_progress = QProgressBar()
        self.sales_progress.setRange(0, 100)
        self.sales_progress.setValue(0)
        self.sales_progress.setFormat("%p%")
        self.sales_progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #4CAF50;
                border-radius: 5px;
                text-align: center;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 3px;
            }
        """)
        sales_progress_layout.addWidget(self.sales_progress)
        sales_progress_group.setLayout(sales_progress_layout)
        progress_layout.addWidget(sales_progress_group)
        
        # Profit progress
        profit_progress_group = QGroupBox("Profit Progress")
        profit_progress_group.setStyleSheet("QGroupBox { font-weight: bold; color: #2196F3; }")
        profit_progress_layout = QVBoxLayout()
        self.profit_progress = QProgressBar()
        self.profit_progress.setRange(0, 100)
        self.profit_progress.setValue(0)
        self.profit_progress.setFormat("%p%")
        self.profit_progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #2196F3;
                border-radius: 5px;
                text-align: center;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #2196F3;
                border-radius: 3px;
            }
        """)
        profit_progress_layout.addWidget(self.profit_progress)
        profit_progress_group.setLayout(profit_progress_layout)
        progress_layout.addWidget(profit_progress_group)
        
        # Time tracking
        time_progress_group = QGroupBox("Current Time")
        time_progress_group.setStyleSheet("QGroupBox { font-weight: bold; color: #FF9800; }")
        time_progress_layout = QVBoxLayout()
        self.time_label = QLabel()
        self.time_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.time_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #FF9800;")
        time_progress_layout.addWidget(self.time_label)
        time_progress_group.setLayout(time_progress_layout)
        progress_layout.addWidget(time_progress_group)
        
        layout.addLayout(progress_layout)
        
        # Recent activity
        recent_group = QGroupBox("Recent Activity")
        recent_group.setStyleSheet("QGroupBox { font-weight: bold; color: #333; }")
        recent_layout = QVBoxLayout()
        self.recent_table = QTableWidget()
        self.recent_table.setColumnCount(4)
        self.recent_table.setHorizontalHeaderLabels(["Invoice No", "Date", "Items", "Profit"])
        self.recent_table.horizontalHeader().setStretchLastSection(True)
        self.recent_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
                selection-background-color: #4CAF50;
            }
            QHeaderView::section {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
        """)
        recent_layout.addWidget(self.recent_table)
        recent_group.setLayout(recent_layout)
        layout.addWidget(recent_group)
        
        self.setLayout(layout)
        
        # Initial update
        self.update_dashboard()
    
    def create_stat_card(self, title, value, color):
        card = QFrame()
        card.setFrameStyle(QFrame.Shape.Box)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: white;
                border: 2px solid {color};
                border-radius: 10px;
                padding: 15px;
                margin: 5px;
            }}
            QFrame:hover {{
                background-color: {color}10;
            }}
        """)
        
        layout = QVBoxLayout()
        
        title_label = QLabel(title)
        title_label.setStyleSheet(f"font-size: 14px; color: {color}; font-weight: bold;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        value_label = QLabel(value)
        value_label.setObjectName("value_label")
        value_label.setStyleSheet(f"font-size: 28px; font-weight: bold; color: {color};")
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(value_label)
        
        card.setLayout(layout)
        return card
    
    def update_dashboard(self):
        cursor = self.parent.conn.cursor()
        
        # Update sales
        cursor.execute("SELECT SUM(selling_price * quantity) FROM invoice_items")
        total_sales = cursor.fetchone()[0] or 0
        self.sales_label.setText(f"${total_sales:.2f}")
        
        # Update profit
        cursor.execute("SELECT SUM(profit) FROM invoice_items")
        total_profit = cursor.fetchone()[0] or 0
        self.profit_label.setText(f"${total_profit:.2f}")
        
        # Update products count
        cursor.execute("SELECT COUNT(*) FROM products")
        total_products = cursor.fetchone()[0] or 0
        self.products_label.setText(str(total_products))
        
        # Update invoices count
        cursor.execute("SELECT COUNT(*) FROM invoices")
        total_invoices = cursor.fetchone()[0] or 0
        self.invoices_label.setText(str(total_invoices))
        
        # Update progress bars
        monthly_target = 10000
        if total_sales > 0:
            self.sales_progress.setValue(min(int((total_sales / monthly_target) * 100), 100))
        if total_profit > 0:
            self.profit_progress.setValue(min(int((total_profit / (monthly_target * 0.2)) * 100), 100))
        
        # Update time
        current_time = QDateTime.currentDateTime()
        self.time_label.setText(current_time.toString("hh:mm AP"))
        
        # Update recent activity
        cursor.execute("""
            SELECT i.invoice_no, i.date, COUNT(ii.id) as items, SUM(ii.profit) as profit
            FROM invoices i
            LEFT JOIN invoice_items ii ON i.id = ii.invoice_id
            GROUP BY i.id
            ORDER BY i.date DESC, i.time DESC
            LIMIT 10
        """)
        recent_data = cursor.fetchall()
        
        self.recent_table.setRowCount(len(recent_data))
        for row, data in enumerate(recent_data):
            for col, value in enumerate(data):
                if col == 3 and value:
                    item = QTableWidgetItem(f"${value:.2f}")
                    if value >= 0:
                        item.setForeground(QColor("#4CAF50"))
                    else:
                        item.setForeground(QColor("#f44336"))
                else:
                    item = QTableWidgetItem(str(value))
                self.recent_table.setItem(row, col, item)